"""Excel レポート出力。集計値は Python 側で確定させたデータを書き、合計行だけ SUM 式にする。"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"


def _autofit(ws, max_width: int = 60):
    for col in ws.columns:
        width = 8
        for c in col:
            if c.value is not None:
                width = max(width, min(max_width, len(str(c.value)) * (2 if _is_wide(str(c.value)) else 1) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def _is_wide(s: str) -> bool:
    return any(ord(ch) > 0x2E7F for ch in s)


def _style(ws, header_fill="DDEBF7"):
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = "#,##0;[Red]-#,##0;-"
    ws.freeze_panes = "A2"


def write_report(path: str | Path, sheets: dict[str, pd.DataFrame], notes: list[str] | None = None) -> Path:
    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        # 先頭に読み方シート
        readme = pd.DataFrame({"メモ": notes or []})
        readme.to_excel(xw, sheet_name="読み方", index=False)
        for name, df in sheets.items():
            if df is None:
                continue
            d = df.copy()
            for c in d.columns:
                if pd.api.types.is_datetime64_any_dtype(d[c]):
                    d[c] = d[c].dt.strftime("%Y-%m-%d")
            d.to_excel(xw, sheet_name=name[:31], index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        _style(ws)
        _autofit(ws)
        # 月次推移: 合計行を SUM 式で追加(値のハードコードを避ける)
        if ws.title.startswith("月次推移") and ws.max_row >= 2:
            last = ws.max_row
            ws.cell(row=last + 1, column=1, value="合計").font = Font(name=FONT, bold=True)
            for col in range(2, ws.max_column + 1):
                letter = get_column_letter(col)
                c = ws.cell(row=last + 1, column=col, value=f"=SUM({letter}2:{letter}{last})")
                c.font = Font(name=FONT, bold=True)
                c.number_format = "#,##0;[Red]-#,##0;-"
    ws0 = wb["読み方"]
    ws0.column_dimensions["A"].width = 110
    for row in ws0.iter_rows(min_row=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
    return path
